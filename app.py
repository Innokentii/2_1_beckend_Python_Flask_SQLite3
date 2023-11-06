from flask import Flask, render_template, request, jsonify, Response
import sqlite3 as SQL
import openpyxl
from static.PaswordGen import pwdGenerator
from sys import path as p
from pathlib import Path
import os
import datetime
import sys
import json
import ast
import pprint
import requests as REQ
import base64
import smtplib

DB_PATH = ""

app = Flask(__name__)

"""_____Работа с базой данных SQLite3_____"""
# создание БД для 4_registration
connect = SQL.connect('SQLite3_base.db')
cursor = connect.cursor()
cursor.execute("""CREATE TABLE IF NOT EXISTS users(
                password text,
                login text,
                name text,
                phone text,
                mail text
                );""")
connect.commit()
connect.close
# создание БД для 5_load_file
connect = SQL.connect('SQLite3_file.db')
cursor = connect.cursor()
cursor.execute("""CREATE TABLE IF NOT EXISTS fileload(
                id INTEGER PRIMARY KEY,
                name TEXT,
                size TEXT,
                date TEXT,
                time TEXT,
                file BLOB
                );""")
connect.commit()
connect.close
"""___________________________________________________"""

"""_____Создание каркаса сайта_____"""
app = Flask(__name__)
@app.route('/') # Главная страница

@app.route('/beckend.html') # Главная страница
def beckend():
    return render_template('beckend.html')

@app.route('/1_iframe_ban') # Запрет использования iframe в сайте
def iframe_ban():
    return render_template('1_iframe_ban.html')

@app.route('/1_iframe_ban_check') # Проверка запрета использования iframe в сайте
def iframe_ban_check():
    return render_template('1_iframe_ban_check.html')

@app.route('/2_questionnaire') # Викторина
def questionnaire():
    return render_template('2_questionnaire.html')

@app.route('/3_list_of_products', methods=["GET", "POST"]) # Работа с Excell (чтение данных с файла excel, сохранение, экспорт в PDF и excell)
def list_of_products():
    abc_list = ['A','B','C','D','E','F','G','H','I']

    # Сохранение данных в Excell
    msg = ""
    if request.method == "POST":
        msg = request.form.get("pwd")
        msg = msg
        msg = msg.split('][')
        for n in range(len(msg)):
            msg[n] = msg[n].strip('[').rstrip(']').split(';|;')
    workbook = openpyxl.load_workbook(filename='static/3_Excell.xlsx')
    sheet = workbook.active

    if msg != "":
        for N in range(len(msg)):
            for n in range(9):
                if n != 6 and n != 8:
                    sheet[f'{abc_list[n]}{N + 3}'] = msg[N][n].replace('.', ',')
        workbook.save(filename='static/3_Excell.xlsx')

    # Вывод данных с Excell
    excell = openpyxl.open("static/3_Excell.xlsx", read_only=True, data_only=True)
    sheet = excell.active
    str_table = ''
    count_row = sheet.max_row
    count_column = sheet.max_column
    for line in range(3, count_row + 1):
        row_table = '['
        for column in range(count_column):
            str_add = sheet[f'{abc_list[column]}{line}'].value
            if column != count_column - 1:
                row_table += str(str_add) + ';|; '
            else:
                row_table += str(str_add) + ']'
        str_table += row_table
    str_table = str_table.strip('[').rstrip(']')

    return render_template('3_list_of_products.html', data_3 = str(count_row), exc_f=str(str_table))

@app.route('/4_registration', methods=["GET","POST"]) # Регистрация (логин и пароль)
def registration():
    permission = ''
    base_users = ''
    password_user = ''

    connect = SQL.connect('SQLite3_base.db')
    cursor = connect.cursor()

    if request.method == "POST":
        # Проверка пароля и логина
        login_open_ = request.form.get('login_open_post')
        password_open_ = request.form.get('password_open_post')
        num_char_open = len(login_open_ + password_open_)*3
        password_user = pwdGenerator(password_open_, login_open_, num_char_open)
        password_user = "'" + password_user

        # формирование пароля
        login = request.form.get('login')
        password_enter = request.form.get('password_enter')
        num_char = len(login + password_enter)*3

        if password_enter != '':
            # Запись в базу данных________________________________
            password_gen = pwdGenerator(password_enter, login, num_char)
            lfp = request.form.get('LFP')
            phone = str(request.form.get('phone'))
            mail = request.form.get('mail')
            # Проверка на наличие аккаунта

            table_pasword = list(cursor.execute('SELECT login FROM users'))
            if login in list(pas[0] for pas in table_pasword):
                permission = 'такой логин уже существует!!!'
            else:
                cursor.execute("INSERT INTO users VALUES (?, ?, ?, ?, ?)", (password_gen, login, lfp, phone, mail))
                connect.commit()
                permission = 'вы зарегистрированы)'

    base_users = cursor.execute('SELECT * FROM users')
    connect.close

    return render_template('4_registration.html', reg_val = [permission, base_users, password_user])

@app.route('/5_load_file', methods=['GET', 'POST']) # загрузка файла (pdf, word, фото, музыка, видео) на базу данных
def load_file():
    connect = SQL.connect('SQLite3_file.db')
    cursor = connect.cursor()
    if request.method == 'POST':
        info_url_py = request.files['info_url_py']
        if info_url_py:

            # Путь к файлу
            file_path = ''
            file_work = info_url_py.filename
            if os.path.isfile(f'static/upload/{info_url_py.filename}') == False:
                file_path = os.path.join('static/upload', info_url_py.filename)
            else:
                count_copy = 0
                while True:
                    count_copy += 1
                    file_work = ((info_url_py.filename)[::-1].replace('.', f'.){str(count_copy)[::-1]}(ypoc_', 1))[::-1]
                    if os.path.isfile(f'static/upload/{file_work}') == False:
                        file_path = Path(p[0]) / f'static/upload/{file_work}'
                        break
            # Сохранение на папку upload
            info_url_py.save(file_path)

            # вывод времени
            new_date = str(datetime.date.today())
            new_time = str(datetime.datetime.now().time().strftime('%H:%M:%S'))

            # СОХРАНЕНИЕ В БАЗУ ДАННЫХ

            # Имя файла
            file_name = file_work

            # Размер файла
            size_file = str(os.stat(f'static/upload/{info_url_py.filename}').st_size)

            # Чтение файла
            with open(f'static/upload/{file_work}', 'rb') as file_save:
                file_path = file_save.read()

            # СОХРАНЕНИЕ В БАЗУ ДАННЫХ
            cursor.execute('INSERT INTO fileload (name, size, date, time, file) VALUES (?, ?, ?, ?, ?)', [file_name, size_file, new_date, new_time, file_path])
            connect.commit()
    connect.close
    
    return render_template('5_load_file.html')

@app.route('/6_discharge_file', methods = ['GET', 'POST']) # скачивание файла (pdf, word, фото, музыка, видео) на рабочий компьютер
def discharge_file():
    _6_info_js = request.form.get('infopy')
    info_files_1 = ''
    info_files_2 = ''
    info_files_3 = ''
    info_files_4 = ''
    count_row = '0'

    connect = SQL.connect('SQLite3_file.db')
    cursor = connect.cursor()
    if request.method == "POST":
        info_files_1 = list(cursor.execute('SELECT name FROM fileload'))
        info_files_2 = list(cursor.execute('SELECT size FROM fileload'))
        info_files_3 = list(cursor.execute('SELECT date FROM fileload'))
        info_files_4 = list(cursor.execute('SELECT time FROM fileload'))
        count_row = str(len(info_files_1))
        if _6_info_js == 'A':
            load_file = list(cursor.execute('SELECT file FROM fileload'))
            info_files_name = list(cursor.execute('SELECT name FROM fileload'))
            k = 0
            for f in load_file:
                file_name = ''
                file_name = str(info_files_name[k])[2:-3]
                if os.path.isfile(f'{Path.home() / "Downloads"}/{file_name}') == True:
                    count_copy = 0
                    while True:
                        count_copy += 1
                        file_name = ((file_name)[::-1].replace('.', f'.){str(count_copy)[::-1]}(ypoc_', 1))[::-1]
                        if os.path.isfile(f'{Path.home() / "Downloads"}/{file_name}') == False:
                            break
                with open(f'{Path.home() / "Downloads"}/{file_name}', 'wb') as file:
                    file.write(f[0])
                    k += 1
        if _6_info_js != '0' and _6_info_js != 'A':
            num = int(_6_info_js)
            load_file = list(cursor.execute('SELECT file FROM fileload'))[num -1 ]
            info_files_name = list(cursor.execute('SELECT name FROM fileload'))[num -1]
            file_name = str(info_files_name)[2:-3]
            if os.path.isfile(f'{Path.home() / "Downloads"}/{file_name}') == True:
                count_copy = 0
                while True:
                    count_copy += 1
                    file_name = ((file_name)[::-1].replace('.', f'.){str(count_copy)[::-1]}(ypoc_', 1))[::-1]
                    if os.path.isfile(f'{Path.home() / "Downloads"}/{file_name}') == False:
                        break
            with open(f'{Path.home() / "Downloads"}/{file_name}', 'wb') as file:
                file.write(load_file[0])

    connect.close

    return render_template('6_discharge_file.html', info_files = [count_row, info_files_1, info_files_2, info_files_3, info_files_4])

@app.route('/7_search_engine') # Статичный и динамический поиск данных SQL
def search_engine():
    return render_template('7_search_engine.html')

@app.route('/7_search_engine_get', methods=['GET']) # Статичный и динамический поиск данных SQL (Передача данных)
def search_engine_get():
    connect = SQL.connect('SQLite3_file.db')
    cursor = connect.cursor()
    search_object = list(cursor.execute('SELECT name, size, date, time FROM fileload'))
    connect.close
    return jsonify(search_object)


get_text = 0
@app.route('/7_search_engine_get_blob_obj', methods=['GET', 'POST']) # Статичный и динамический поиск данных SQL (Передача данных)
def search_engine_get_blob_obj():
    global get_text
    search_object = ''
    if request.method == 'POST':
        get_text = request.get_json(force=True)
        get_text = int(get_text['param1'])

    connect = SQL.connect('SQLite3_file.db')
    cursor = connect.cursor()
    cursor.execute('SELECT file FROM fileload')
    search_object = [row[0] for row in cursor.fetchall()]
    connect.close
    return Response(search_object[get_text], content_type='application/octet-stream')

@app.route('/8_audio_video') # Запуск аудио в сайте
def audio_video():
    
    return render_template('8_audio_video.html')

get_text_2 = 0
@app.route('/8_audio_video_post', methods=['POST']) # Запуск аудио в сайте (GET)
def audio_video_get():
    global get_text_2
    req = request.get_json(force=True)
    get_text_2 = int(req['param2'])
    return req

    
@app.route('/8_audio_video_get', methods=['GET']) # Запуск аудио в сайте (GET_3)
def audio_video_get_3():
    connect = SQL.connect('SQLite3_file.db')
    cursor = connect.cursor()
    cursor.execute('SELECT name, file FROM fileload')
    connect.close
    file = [row[1] for row in cursor.fetchall() if row[0][-4:]=='.mp3']
    return Response(file[get_text_2], content_type='application/octet-stream')

@app.route('/8_audio_video_get_2', methods=['GET']) # Запуск аудио в сайте (GET_2)
def audio_video_get_2():
    connect = SQL.connect('SQLite3_file.db')
    cursor = connect.cursor()
    files = list(cursor.execute('SELECT name FROM fileload'))
    files = [row[0] for row in files if row[0][-4:]=='.mp3']
    connect.close
    return jsonify(files)

get_text_3 = 0
@app.route('/8_audio_video_post_4', methods=['POST']) # Запуск аудио в сайте (GET_4)
def audio_video_post_4():
    global get_text_3
    # POST запрос
    req = request.get_json(force=True)
    get_text_3 = int(req['param3'])
    return req


@app.route('/8_audio_video_get_4', methods=['GET']) # Запуск аудио в сайте (GET_4)
def audio_video_get_4():
    # GET запрос
    connect = SQL.connect('SQLite3_file.db')
    cursor = connect.cursor()
    files = list(cursor.execute('SELECT name FROM fileload'))
    files = [row[0] for row in files if row[0][-4:]=='.mp3']
    connect.close
    return jsonify(files[get_text_3])

@app.route('/9_video') # Запуск видео в сайте
def video_f():
    return render_template('9_video.html')


@app.route('/9_video_get_1_file', methods = ['GET']) # Запуск видео в сайте (GET запрос file)
def video_get_1_file_f():
    connect = SQL.connect('SQLite3_file.db')
    cursor = connect.cursor()
    cursor.execute('SELECT name FROM fileload')
    files = [row[0] for row in cursor.fetchall() if row[0][-4:] == '.mp4']

    connect.close
    return jsonify(files)

num_video = 0
@app.route('/9_video_get_post_1_file', methods = ['GET', 'POST']) # Запуск видео в сайте (GET запрос file)
def video_get_post_1_file_f():
    global num_video
    file = ''
    if request.method == 'POST':
        req = request.get_json(force=True)
        num_video = int(req['param'])
        return req
    if request.method == 'GET':
        connect = SQL.connect('SQLite3_file.db')
        cursor = connect.cursor()
        cursor.execute('SELECT name, file FROM fileload')
        file = list(row[1] for row in cursor.fetchall() if row[0][-4:] == '.mp4')
        connect.close
        return Response(file[num_video], content_type='application/octet-stream')
    
@app.route('/10_weather_forecast') # API прогноз погоды;
def weather_f():
    return render_template('10_weather_forecast.html')

"""_____ Функция запуска _____"""
def create_app():
    """Функция запуска"""
    return app
